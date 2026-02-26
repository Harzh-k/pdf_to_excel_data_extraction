#!/usr/bin/env python3
"""
extract_tables_smart_merged.py — Universal IRDAI PDF → Excel Extractor
=======================================================================
Hybrid engine: automatically picks the right extraction strategy per page.

  Engine 1 — pdfplumber (lines)    : PDFs with vertical grid lines (HDFC, LIC)
  Engine 2 — gap-based (h_only)    : PDFs with horizontal rules only (Aditya Birla)
  Engine 3 — docling (no_lines)    : Scanned or whitespace-only PDFs

Usage:
    python extract_tables_smart_merged.py <pdf_path> [output_path]
    python extract_tables_smart_merged.py <pdf_path> [output_path] --force-docling
"""

import re
import sys
import os
import logging
from pathlib import Path

try:
    import pdfplumber
    from openpyxl import Workbook
    from openpyxl.styles import Font, Alignment, Border, Side, PatternFill
    from openpyxl.utils import get_column_letter
except ImportError as e:
    print(f"Missing dependency: {e}\nInstall: pip install pdfplumber openpyxl")
    sys.exit(1)

from src.extractor          import TABLE_SETTINGS, process_table
from src.extractor_gap      import extract_gap_based, should_use_header_cols, extract_index_page, detect_index_page
from src.pdf_type_detector  import classify_page

logging.basicConfig(
    level=logging.INFO,
    format="%(asctime)s  %(levelname)-7s  %(message)s",
    datefmt="%H:%M:%S",
)
logger = logging.getLogger(__name__)


# ══════════════════════════════════════════════════════════════════════════════
# STYLES  (unchanged from previous version)
# ══════════════════════════════════════════════════════════════════════════════

_DB    = "1F4E79"
_S_T   = Side(style="thin",   color="BFBFBF")
_S_M   = Side(style="medium", color=_DB)
B_NONE = Border()
B_HDR  = Border(top=_S_M,  bottom=_S_M)
B_TOT  = Border(top=_S_T,  bottom=_S_T)
B_SHR  = Border(bottom=_S_M)
F_DB   = PatternFill("solid", fgColor=_DB)
F_NO   = PatternFill("none")
FT_SH  = Font(name="Calibri", bold=True,  size=10, color="FFFFFF")
FT_CH  = Font(name="Calibri", bold=True,  size=9,  color="FFFFFF")
FT_B   = Font(name="Calibri", bold=True,  size=9)
FT_N   = Font(name="Calibri", bold=False, size=9)
FT_M   = Font(name="Calibri", bold=False, size=8,  color="595959")
A_WR   = Alignment(vertical="top",    wrap_text=True)
A_CT   = Alignment(vertical="center", horizontal="center", wrap_text=True)
A_ML   = Alignment(vertical="center", indent=1)
A_NR   = Alignment(vertical="top",    horizontal="right")

_COL_KW  = {"LIFE","PENSION","HEALTH","ANNUITY","LINKED BUSINESS",
            "PARTICIPATING","NON-PARTICIPATING","VAR.INS","VAR. INS","PARTICULARS",
            "FORM NO","SR NO","FORM NO.","SR NO.","PAGE NO","PAGE NO."}
_TOT_KW  = {"TOTAL (A)","TOTAL (B)","TOTAL (C)","TOTAL (D)",
            "SUB TOTAL","SUB-TOTAL","SURPLUS","DEFICIT","AMOUNT AVAILABLE","TOTAL"}
_FORM_KW = {"FORM L-","REVENUE ACCOUNT","PROFIT AND LOSS",
            "BALANCE SHEET","POLICYHOLDERS","NAME OF THE INSURER"}

# Matches L-XX form codes like L-26, L-1-A-RA, L-14A etc.
_FORM_CODE_RE = re.compile(r"^L-[\dA-Za-z]")


def _is_index_data_row(row):
    """
    True when row is an index/TOC entry: [number, L-XX code, description...].
    These must NEVER be styled blue even if description contains "Life"/"Linked".
    Pattern: first cell is a digit, second cell starts with L-
    """
    cells = [str(c).strip() for c in row if c and str(c).strip()]
    if len(cells) < 2:
        return False
    return cells[0].replace(".", "").isdigit() and _FORM_CODE_RE.match(cells[1])


def classify_row(row, idx, is_idx=False):
    """
    Classify a table row for Excel styling.
    Returns: col_header | total | form_meta | section | normal

    col_header  → dark blue fill, white bold font  (column header rows)
    total       → bold, thin border top+bottom     (subtotal / total rows)
    form_meta   → small grey font                  (form title / insurer info)
    section     → bold                             (section heading rows)
    normal      → regular                          (data rows)
    """
    # Index/TOC pages: only the very first row is a header
    if is_idx:
        return "col_header" if idx == 0 else "normal"

    # Index data rows (number | L-XX | description): never miscolor as header
    if _is_index_data_row(row):
        return "normal"

    t = " ".join(str(c) for c in row if c).upper().strip()

    # col_header guard: row must contain IRDAI column keywords
    # AND must have few actual data numbers (headers have col names, not values)
    if any(k in t for k in _COL_KW):
        num_vals = sum(
            1 for c in row
            if c and re.match(r"^-?\d[\d,\.]+$", str(c).strip())
        )
        if num_vals <= 2:          # allow "-" dashes but not actual figures
            return "col_header"

    # Total rows
    for k in _TOT_KW:
        if re.search(r"\b" + re.escape(k) + r"\b", t):
            return "total"

    # Form/insurer metadata (first few rows of every sheet)
    if idx < 6 and any(k in t for k in _FORM_KW):
        return "form_meta"

    # All-caps multi-word section headings
    ws = [w for w in t.split() if len(w) > 2]
    if ws and all(w.isupper() for w in ws) and len(ws) >= 2:
        return "section"

    return "normal"


def is_num(v):
    if not v: return False
    try: float(str(v).replace(",","").replace("(","").replace(")","").replace("-","").strip()); return True
    except: return False


def apply_style(ws, er, row, rt, nc):
    for ci, val in enumerate(row, 1):
        c = ws.cell(row=er, column=ci, value=val if val else None)
        n = ci > 1 and is_num(val)
        if   rt == "col_header": c.font=FT_CH; c.fill=F_DB; c.border=B_HDR; c.alignment=A_CT
        elif rt == "total":      c.font=FT_B;  c.fill=F_NO; c.border=B_TOT; c.alignment=A_NR if n else A_WR
        elif rt == "form_meta":  c.font=FT_M;  c.fill=F_NO; c.border=B_NONE; c.alignment=A_WR
        elif rt == "section":    c.font=FT_B;  c.fill=F_NO; c.border=B_NONE; c.alignment=A_WR
        else:                    c.font=FT_N;  c.fill=F_NO; c.border=B_NONE; c.alignment=A_NR if n else A_WR


def set_col_widths(ws):
    w = {}
    for row in ws.iter_rows():
        for c in row:
            if c.value:
                w[c.column] = max(w.get(c.column, 8), len(str(c.value).split("\n")[0]))
    for col, width in w.items():
        ws.column_dimensions[get_column_letter(col)].width = min(width + 2, 40)


def make_sheet_name(pb, used):
    b = f"P{pb['page_number']}"
    if pb.get("form_name"):
        b = f"P{pb['page_number']}_{pb['form_name'].replace('FORM ','').strip()}"
    b = re.sub(r"[\\/*?:\[\]]", "", b)[:31]
    if b not in used: used.add(b); return b
    for n in range(2, 100):
        c = f"{b[:28]}_{n}"
        if c not in used: used.add(c); return c
    return b


# ══════════════════════════════════════════════════════════════════════════════
# HYBRID EXTRACTION ENGINE
# ══════════════════════════════════════════════════════════════════════════════

def _extract_page_pdfplumber(page, current_form):
    """
    Engine 1: pdfplumber line-based extraction with smart fallback.

    Priority:
      0. index_page  — TOC/index pages (Sl.No / Form No / Description / Pages)
                       pdfplumber collapses all 3 left cols into one cell.
      1. pdfplumber  — run first, always.
      2. header_cols — override pdfplumber when it under-detects columns OR
                       merges a data value into the Particulars column (col-0).
                       Condition checked by should_use_header_cols().
      3. gap-based   — fallback when gap scores >1.5x pdfplumber AND has >= cols.
    """
    # ── Priority 0: Index/TOC page ────────────────────────────────────────────
    words = page.extract_words(x_tolerance=2, y_tolerance=4)
    if detect_index_page(words):
        idx_data = extract_index_page(page)
        if idx_data:
            logger.info(
                f"  Page {page.page_number} [index_page]: "
                f"{len(idx_data)}r x {len(idx_data[0]) if idx_data else 0}c"
            )
            return [{"type": "table", "data": idx_data}]

    text   = page.extract_text() or ""
    tables = page.find_tables(table_settings=TABLE_SETTINGS)

    # ── Run pdfplumber first ──────────────────────────────────────────────────
    pp_rows = sum(len(t.extract()) for t in tables) if tables else 0
    pp_cols = max((len(t.extract()[0]) for t in tables if t.extract()), default=1) if tables else 1
    pp_score = pp_rows * pp_cols

    # Build the rebuilt pdfplumber table (needed for col-0 merge check)
    pp_data = []
    if tables:
        pp_data = process_table(page, tables[0])

    # ── Smart check: should we use header_cols instead? ───────────────────────
    use_hdr, hdr_data = should_use_header_cols(page, pp_data)
    if use_hdr and hdr_data:
        return [{"type": "table", "data": hdr_data}]

    # ── gap-based fallback (only when clearly better) ─────────────────────────
    gap_data = extract_gap_based(page)
    gap_rows = len(gap_data) if gap_data else 0
    gap_cols = len(gap_data[0]) if gap_data else 0
    gap_score = gap_rows * gap_cols

    # ── NEW: detect spurious empty columns in pdfplumber output ───────────────
    # When pdfplumber detects phantom vertical lines (e.g. a thin guide column
    # at the left margin), it generates columns that are 100% empty in data rows.
    # In that case, prefer gap-based if it produces a cleaner result.
    def _pp_has_empty_cols(data, skip_rows=5):
        """Return True if pdfplumber output has any 100%-empty middle column."""
        if not data or len(data[0]) < 3:
            return False
        data_rows = data[skip_rows:] if len(data) > skip_rows else data
        if not data_rows:
            return False
        ncols = len(data[0])
        for ci in range(1, ncols - 1):  # skip first and last col
            nonempty = sum(1 for row in data_rows if ci < len(row) and row[ci] not in ('', None))
            if nonempty == 0:
                return True
        return False

    pp_has_phantom = pp_data and _pp_has_empty_cols(pp_data)

    if (gap_data and gap_score > pp_score * 1.5 and gap_cols >= pp_cols):
        logger.info(
            f"  Page {page.page_number} [gap>pdfplumber {gap_score}>{pp_score}]: "
            f"{gap_rows}r x {gap_cols}c"
        )
        return [{"type": "table", "data": gap_data}]

    # If pp has phantom empty columns and gap gives a simpler result, use gap
    # Only apply this for small tables (pp_cols <= 8). Wide tables (20+ cols)
    # may have structural empty columns that are legitimate, and gap often under-extracts them.
    if pp_has_phantom and gap_data and gap_cols >= 2 and pp_cols <= 8:
        logger.info(
            f"  Page {page.page_number} [gap>pp (phantom cols) {gap_cols}c<{pp_cols}c]: "
            f"{gap_rows}r x {gap_cols}c"
        )
        return [{"type": "table", "data": gap_data}]

    # ── Use pdfplumber result ─────────────────────────────────────────────────
    content = []
    if tables:
        for t in tables:
            fin = process_table(page, t)
            content.append({"type": "table", "data": fin})
            logger.info(
                f"  Page {page.page_number} [pdfplumber]: "
                f"{len(fin)}r x {len(fin[0]) if fin else 0}c"
            )
    else:
        lines = [l for l in text.split("\n") if l.strip()]
        if lines:
            content.append({"type": "text", "data": lines})

    return content


def _extract_page_gap(page, current_form):
    """Engine 2: gap-based for h_only PDFs. Tries header_cols when applicable."""
    content = []

    # ── Priority 0: Index/TOC page (works for both lines and h_only pages) ───
    words = page.extract_words(x_tolerance=2, y_tolerance=4)
    if detect_index_page(words):
        idx_data = extract_index_page(page)
        if idx_data:
            logger.info(
                f"  Page {page.page_number} [index_page]: "
                f"{len(idx_data)}r x {len(idx_data[0]) if idx_data else 0}c"
            )
            return [{"type": "table", "data": idx_data}]

    # For h_only pages that ARE segmental forms (many col headers), use header_cols
    use_hdr, hdr_data = should_use_header_cols(page, None)
    if use_hdr and hdr_data:
        return [{"type": "table", "data": hdr_data}]

    rows = extract_gap_based(page)
    if rows:
        content.append({"type": "table", "data": rows})
        logger.info(
            f"  Page {page.page_number} [gap-based]: "
            f"{len(rows)}r x {len(rows[0]) if rows else 0}c"
        )
    else:
        text  = page.extract_text() or ""
        lines = [l for l in text.split("\n") if l.strip()]
        if lines:
            content.append({"type": "text", "data": lines})
        logger.warning(f"  Page {page.page_number} [gap-based]: failed, using text")

    return content


def extract_document(pdf_path: str, force_docling: bool = False) -> list:
    """
    Main extraction entry point.
    Automatically routes each page to the right engine.

    Args:
        pdf_path:      path to PDF file
        force_docling: skip detection, use docling for ALL pages

    Returns: list of page_block dicts
    """
    document    = []
    current_form = None

    # ── Check if docling is needed ────────────────────────────────────────────
    docling_pages = []

    with pdfplumber.open(pdf_path) as pdf:
        total = len(pdf.pages)
        logger.info(f"Opened '{Path(pdf_path).name}' — {total} pages")

        for page in pdf.pages:
            pg_num  = page.page_number
            text    = page.extract_text() or ""

            # Track form name
            m = re.search(r"FORM\s+L-[\dA-Za-z\-]+", text)
            if m: current_form = m.group().strip()

            pb = {
                "page_number": pg_num,
                "form_name":   current_form,
                "is_index":    False,
                "content":     [],
                "engine":      None,
            }

            if force_docling:
                docling_pages.append(pg_num)
                document.append(pb)
                continue

            # ── Detect page type and route ────────────────────────────────────
            pg_type = classify_page(page)
            pb["engine"] = pg_type

            if pg_type == "lines":
                pb["content"] = _extract_page_pdfplumber(page, current_form)

            elif pg_type == "h_only":
                pb["content"] = _extract_page_gap(page, current_form)

            else:  # 'no_lines' or 'scanned'
                docling_pages.append(pg_num)
                logger.info(f"  Page {pg_num} [{pg_type}]: queued for docling")

            # Detect index page
            for item in pb["content"]:
                if item["type"] == "table":
                    raw = item["data"]
                    if (len(raw) == 2 and raw[1] and
                            any("\n" in str(c) for c in raw[1] if c)):
                        pb["is_index"] = True

            document.append(pb)

    # ── Docling pass for no_lines / scanned pages ─────────────────────────────
    if docling_pages:
        logger.info(
            f"\n{len(docling_pages)} pages need docling "
            f"(no_lines/scanned): {docling_pages[:10]}{'...' if len(docling_pages)>10 else ''}"
        )
        try:
            from src.extractor_docling import extract_with_docling
            docling_blocks = extract_with_docling(pdf_path, page_numbers=docling_pages)

            # Merge docling results back into document
            docling_map = {b["page_number"]: b for b in docling_blocks}
            for pb in document:
                if pb["page_number"] in docling_map:
                    db = docling_map[pb["page_number"]]
                    pb["content"] = db["content"]
                    pb["engine"]  = "docling"

        except ImportError:
            logger.warning(
                "docling not installed — pages with no lines will be extracted "
                "as plain text.\n"
                "Install with: pip install docling"
            )
            # Fallback: try gap-based on those pages too
            with pdfplumber.open(pdf_path) as pdf:
                for page in pdf.pages:
                    if page.page_number not in docling_pages:
                        continue
                    for pb in document:
                        if pb["page_number"] == page.page_number:
                            pb["content"] = _extract_page_gap(page, pb.get("form_name"))
                            pb["engine"]  = "gap-fallback"

    return document


# ══════════════════════════════════════════════════════════════════════════════
# EXCEL WRITER  (unchanged)
# ══════════════════════════════════════════════════════════════════════════════

def write_excel(document: list, output_path=None):
    """
    Build and save workbook.
    output_path: str  -> save to disk AND return bytes
    output_path: None -> return bytes only (used by Streamlit)
    """
    import io as _io
    wb, used = Workbook(), set()
    wb.remove(wb.active)

    for pb in document:
        sn = make_sheet_name(pb, used)
        ws = wb.create_sheet(title=sn)
        cr = 1

        title = f"Page {pb['page_number']}"
        if pb.get("form_name"): title += f"   |   {pb['form_name']}"
        hc = ws.cell(row=1, column=1, value=title)
        hc.font=FT_SH; hc.fill=F_DB; hc.border=B_SHR; hc.alignment=A_ML
        cr = 2

        for item in pb["content"]:
            if item["type"] == "table":
                td = item["data"]
                if not td: continue
                nc = max(len(r) for r in td)
                for ri, row in enumerate(td):
                    row = list(row) + [""] * (nc - len(row))
                    if not any(str(v).strip() for v in row):
                        continue
                    rt  = classify_row(row, ri, pb.get("is_index", False))
                    apply_style(ws, cr, row, rt, nc)
                    cr += 1
                cr += 1
            else:
                for line in item["data"]:
                    c = ws.cell(row=cr, column=1, value=line)
                    c.font=FT_N; c.alignment=A_WR
                    cr += 1
                cr += 1

        set_col_widths(ws)
        ws.freeze_panes = "A2"

    buf = _io.BytesIO()
    wb.save(buf)
    raw = buf.getvalue()

    if output_path and isinstance(output_path, str):
        try:
            with open(output_path, "wb") as f:
                f.write(raw)
            size_kb = len(raw) // 1024
            logger.info(f"Saved '{output_path}' ({size_kb} KB, {len(document)} sheets)")
        except PermissionError:
            logger.error(f"File is open — close it first: {output_path}")
            sys.exit(1)

    return raw


# ══════════════════════════════════════════════════════════════════════════════
# MAIN
# ══════════════════════════════════════════════════════════════════════════════

def main():
    if len(sys.argv) < 2:
        print(__doc__)
        sys.exit(1)

    pdf_path      = sys.argv[1]
    force_docling = "--force-docling" in sys.argv

    if not os.path.exists(pdf_path):
        logger.error(f"File not found: {pdf_path}")
        sys.exit(1)

    output_path = (
        sys.argv[2] if len(sys.argv) >= 3 and not sys.argv[2].startswith("--")
        else str(Path(pdf_path).with_suffix("")) + "_IRDAI.xlsx"
    )
    if not output_path.lower().endswith(".xlsx"):
        output_path += ".xlsx"

    logger.info("=" * 60)
    logger.info("IRDAI Hybrid PDF → Excel Extractor")
    logger.info("=" * 60)

    document = extract_document(pdf_path, force_docling=force_docling)
    if not document:
        logger.warning("No content extracted.")
        sys.exit(0)

    write_excel(document, output_path)

    logger.info("=" * 60)
    logger.info("Done.")
    logger.info("=" * 60)


if __name__ == "__main__":
    main()