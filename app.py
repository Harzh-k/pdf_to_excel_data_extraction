"""
app.py — IRDAI PDF → Excel Extractor  |  Streamlit UI
Run:  streamlit run app.py
"""

import io
import os
import re
import time
import logging
import tempfile
from pathlib import Path

import streamlit as st

st.set_page_config(
    page_title="IRDAI PDF Extractor",
    page_icon="📊",
    layout="centered",
    initial_sidebar_state="collapsed",
)

st.markdown("""
<style>
/* ── Reset & base ─────────────────────────────────────── */
#MainMenu, footer, header { visibility: hidden; }
.stApp { background: #F8FAFC; }

/* ── Page wrapper — max width centered ───────────────── */
.block-container {
    max-width: 780px !important;
    padding: 2.5rem 1.5rem 3rem !important;
}

/* ── Hero ────────────────────────────────────────────── */
.hero {
    text-align: center;
    padding: 40px 0 32px;
}
.hero-icon  { font-size: 3rem; line-height: 1; margin-bottom: 12px; }
.hero-title {
    font-size: 1.9rem; font-weight: 800;
    color: #0F2942; margin: 0 0 8px;
    letter-spacing: -0.5px;
}
.hero-sub {
    font-size: 0.95rem; color: #64748B;
    margin: 0; line-height: 1.6;
}

/* ── Divider ─────────────────────────────────────────── */
.divider {
    border: none; border-top: 1px solid #E2E8F0;
    margin: 24px 0;
}

/* ── Section label ───────────────────────────────────── */
.section-label {
    font-size: 0.68rem; font-weight: 700;
    letter-spacing: 0.12em; text-transform: uppercase;
    color: #94A3B8; margin: 0 0 8px;
}

/* ── Upload zone — override Streamlit dark uploader ─── */
[data-testid="stFileUploader"] {
    background: #FFFFFF !important;
    border: 2px dashed #CBD5E1 !important;
    border-radius: 12px !important;
    padding: 8px !important;
    transition: border-color 0.2s;
}
[data-testid="stFileUploader"]:hover {
    border-color: #1F4E79 !important;
}
/* Inner drag area — force white */
[data-testid="stFileUploader"] > div {
    background: #FFFFFF !important;
    border-radius: 10px !important;
}
/* "Drag and drop" text */
[data-testid="stFileUploader"] span {
    color: #475569 !important;
}
/* "Limit" sub-text */
[data-testid="stFileUploader"] small {
    color: #94A3B8 !important;
}
/* Browse button — light style */
[data-testid="stFileUploader"] button {
    background: #F1F5F9 !important;
    color: #1F4E79 !important;
    border: 1px solid #CBD5E1 !important;
    border-radius: 6px !important;
    font-weight: 600 !important;
    font-size: 0.85rem !important;
    padding: 6px 18px !important;
    transition: all 0.15s !important;
}
[data-testid="stFileUploader"] button:hover {
    background: #DBEAFE !important;
    border-color: #1F4E79 !important;
    color: #1F4E79 !important;
}
/* File icon area */
[data-testid="stFileUploaderDropzoneInstructions"] {
    color: #64748B !important;
}

/* ── File info pill ──────────────────────────────────── */
.file-pill {
    display: flex; align-items: center; gap: 12px;
    background: #EFF6FF;
    border: 1px solid #BFDBFE;
    border-radius: 10px;
    padding: 12px 16px;
    margin-top: 12px;
}
.file-pill-icon { font-size: 1.4rem; flex-shrink: 0; }
.file-pill-name { font-weight: 700; color: #1E40AF; font-size: 0.9rem; margin: 0; }
.file-pill-size { color: #64748B; font-size: 0.78rem; margin: 2px 0 0; }

/* ── Primary button ──────────────────────────────────── */
.stButton > button[kind="primary"] {
    background: #1F4E79 !important;
    color: #FFFFFF !important;
    border: none !important;
    border-radius: 8px !important;
    font-weight: 700 !important;
    font-size: 1rem !important;
    padding: 12px 0 !important;
    width: 100%;
    letter-spacing: 0.01em;
    box-shadow: 0 2px 8px rgba(31,78,121,0.25);
    transition: all 0.15s;
}
.stButton > button[kind="primary"]:hover {
    background: #163B5E !important;
    box-shadow: 0 4px 12px rgba(31,78,121,0.35);
}

/* ── Progress bar ────────────────────────────────────── */
.stProgress > div > div > div { background: #1F4E79 !important; }

/* ── Stats row ───────────────────────────────────────── */
.stats-row {
    display: flex; gap: 10px; margin: 16px 0;
}
.stat-card {
    flex: 1;
    background: #FFFFFF;
    border: 1px solid #E2E8F0;
    border-radius: 10px;
    padding: 16px 12px;
    text-align: center;
    box-shadow: 0 1px 3px rgba(0,0,0,0.04);
}
.stat-num  { font-size: 1.8rem; font-weight: 800; color: #1F4E79; line-height: 1; }
.stat-lbl  {
    font-size: 0.68rem; font-weight: 700;
    letter-spacing: 0.1em; text-transform: uppercase;
    color: #94A3B8; margin-top: 4px;
}

/* ── Badges ──────────────────────────────────────────── */
.badges { margin: 4px 0 16px; }
.badge {
    display: inline-block;
    background: #EFF6FF; color: #1D4ED8;
    border: 1px solid #BFDBFE;
    border-radius: 4px; font-size: 0.7rem;
    font-weight: 600; padding: 2px 8px;
    margin: 2px 3px 2px 0;
}

/* ── Log box ─────────────────────────────────────────── */
.log-box {
    background: #0F172A;
    border-radius: 8px;
    padding: 14px 16px;
    max-height: 200px;
    overflow-y: auto;
    font-family: 'Courier New', monospace;
    font-size: 0.75rem;
    line-height: 1.7;
    margin-top: 8px;
}
.log-info  { color: #60A5FA; }
.log-ok    { color: #34D399; }
.log-warn  { color: #FBBF24; }
.log-error { color: #F87171; }

/* ── Download button ─────────────────────────────────── */
.stDownloadButton > button {
    background: #16A34A !important;
    color: #FFFFFF !important;
    border: none !important;
    border-radius: 8px !important;
    font-weight: 700 !important;
    font-size: 1rem !important;
    padding: 12px 0 !important;
    width: 100%;
    box-shadow: 0 2px 8px rgba(22,163,74,0.25);
}
.stDownloadButton > button:hover {
    background: #15803D !important;
}

/* ── Text input ──────────────────────────────────────── */
.stTextInput input {
    border-radius: 7px !important;
    border: 1px solid #CBD5E1 !important;
    font-size: 0.88rem !important;
    background: #FFFFFF !important;
    color: #0F172A !important;
}

/* ── Success / info boxes ────────────────────────────── */
[data-testid="stAlert"] {
    border-radius: 8px !important;
}

/* ── Footer ──────────────────────────────────────────── */
.footer {
    text-align: center;
    color: #CBD5E1;
    font-size: 0.72rem;
    margin-top: 40px;
    padding-top: 16px;
    border-top: 1px solid #E2E8F0;
}
</style>
""", unsafe_allow_html=True)


# ══════════════════════════════════════════════════════════════════════════════
# LOGGING
# ══════════════════════════════════════════════════════════════════════════════

class _UILogHandler(logging.Handler):
    def emit(self, record):
        msg = self.format(record)
        if "logs" not in st.session_state:
            st.session_state["logs"] = []
        st.session_state["logs"].append((record.levelname, msg))


def _setup_logging():
    root = logging.getLogger()
    root.setLevel(logging.INFO)
    if not any(isinstance(h, _UILogHandler) for h in root.handlers):
        h = _UILogHandler()
        h.setFormatter(logging.Formatter("%(asctime)s  %(message)s", "%H:%M:%S"))
        root.addHandler(h)


def _render_logs():
    logs = st.session_state.get("logs", [])
    if not logs:
        return
    css_map = {"INFO": "log-info", "WARNING": "log-warn",
               "ERROR": "log-error"}
    lines = [
        f'<span class="{css_map.get(lvl, "log-ok")}">'
        f'{msg.replace("<","&lt;").replace(">","&gt;")}</span>'
        for lvl, msg in logs
    ]
    st.markdown(
        '<div class="log-box">' + "<br>".join(lines) + "</div>",
        unsafe_allow_html=True,
    )


# ══════════════════════════════════════════════════════════════════════════════
# EXTRACTION  (cached per PDF content)
# ══════════════════════════════════════════════════════════════════════════════

@st.cache_data(show_spinner=False)
def _run_extraction(pdf_bytes: bytes):
    import pdfplumber
    from openpyxl import Workbook
    from openpyxl.styles import Font, Alignment, Border, Side, PatternFill
    from openpyxl.utils import get_column_letter
    from src.extractor         import TABLE_SETTINGS, process_table
    from src.extractor_gap     import extract_gap_based
    from src.pdf_type_detector import classify_page

    # Styles
    _DB = "1F4E79"
    S_T = Side(style="thin",   color="BFBFBF")
    S_M = Side(style="medium", color=_DB)
    BN  = Border()
    BH  = Border(top=S_M,  bottom=S_M)
    BT  = Border(top=S_T,  bottom=S_T)
    BSH = Border(bottom=S_M)
    FDB = PatternFill("solid", fgColor=_DB)
    FNO = PatternFill("none")
    FSH = Font(name="Calibri", bold=True,  size=10, color="FFFFFF")
    FCH = Font(name="Calibri", bold=True,  size=9,  color="FFFFFF")
    FB  = Font(name="Calibri", bold=True,  size=9)
    FN  = Font(name="Calibri", bold=False, size=9)
    FM  = Font(name="Calibri", bold=False, size=8,  color="595959")
    AW  = Alignment(vertical="top",    wrap_text=True)
    AC  = Alignment(vertical="center", horizontal="center", wrap_text=True)
    AM  = Alignment(vertical="center", indent=1)
    ANR = Alignment(vertical="top",    horizontal="right")

    COL_KW  = {"LIFE","PENSION","HEALTH","ANNUITY","LINKED BUSINESS",
               "PARTICIPATING","NON-PARTICIPATING","VAR.INS","VAR. INS","PARTICULARS"}
    TOT_KW  = {"TOTAL (A)","TOTAL (B)","TOTAL (C)","TOTAL (D)",
               "SUB TOTAL","SUB-TOTAL","SURPLUS","DEFICIT","AMOUNT AVAILABLE","TOTAL"}
    FORM_KW = {"FORM L-","REVENUE ACCOUNT","PROFIT AND LOSS",
               "BALANCE SHEET","POLICYHOLDERS","NAME OF THE INSURER"}

    def classify(row, idx, is_idx):
        if is_idx: return "col_header" if idx == 0 else "normal"
        t = " ".join(str(c) for c in row if c).upper().strip()
        if any(k in t for k in COL_KW): return "col_header"
        for k in TOT_KW:
            if re.search(r"\b" + re.escape(k) + r"\b", t): return "total"
        if idx < 6 and any(k in t for k in FORM_KW): return "form_meta"
        ws = [w for w in t.split() if len(w) > 2]
        if ws and all(w.isupper() for w in ws) and len(ws) >= 2: return "section"
        return "normal"

    def is_num(v):
        if not v: return False
        try: float(str(v).replace(",","").replace("(","").replace(")","").replace("-","").strip()); return True
        except: return False

    def srow(ws, er, row, rt, nc):
        for ci, val in enumerate(row, 1):
            c = ws.cell(row=er, column=ci, value=val if val else None)
            n = ci > 1 and is_num(val)
            if   rt == "col_header": c.font=FCH; c.fill=FDB; c.border=BH; c.alignment=AC
            elif rt == "total":      c.font=FB;  c.fill=FNO; c.border=BT; c.alignment=ANR if n else AW
            elif rt == "form_meta":  c.font=FM;  c.fill=FNO; c.border=BN; c.alignment=AW
            elif rt == "section":    c.font=FB;  c.fill=FNO; c.border=BN; c.alignment=AW
            else:                    c.font=FN;  c.fill=FNO; c.border=BN; c.alignment=ANR if n else AW

    def set_widths(ws):
        w = {}
        for row in ws.iter_rows():
            for c in row:
                if c.value:
                    w[c.column] = max(w.get(c.column, 8), len(str(c.value).split("\n")[0]))
        for col, width in w.items():
            ws.column_dimensions[get_column_letter(col)].width = min(width + 2, 40)

    def sname(pb, used):
        b = f"P{pb['page_number']}"
        if pb["form_name"]:
            b = f"P{pb['page_number']}_{pb['form_name'].replace('FORM ','').strip()}"
        b = re.sub(r"[\\/*?:\[\]]", "", b)[:31]
        if b not in used: used.add(b); return b
        for n in range(2, 100):
            c = f"{b[:28]}_{n}"
            if c not in used: used.add(c); return c
        return b

    with tempfile.NamedTemporaryFile(suffix=".pdf", delete=False) as tmp:
        tmp.write(pdf_bytes)
        tmp_path = tmp.name

    try:
        document, cur_form = [], None
        stats = {"pages": 0, "tables": 0, "sheets": 0, "forms": set(), "engines": set()}

        with pdfplumber.open(tmp_path) as pdf:
            stats["pages"] = len(pdf.pages)
            for page in pdf.pages:
                text = page.extract_text() or ""
                m = re.search(r"FORM\s+L-[\dA-Za-z\-]+", text)
                if m: cur_form = m.group().strip(); stats["forms"].add(cur_form)

                pb = {"page_number": page.page_number, "form_name": cur_form,
                      "is_index": False, "content": []}

                pg_type = classify_page(page)
                pb["engine"] = pg_type

                if pg_type == "lines":
                    tables = page.find_tables(table_settings=TABLE_SETTINGS)
                    if tables:
                        for t in tables:
                            raw    = t.extract()
                            is_idx = (len(raw)==2 and raw[1] and
                                      any("\n" in str(c) for c in raw[1] if c))
                            if is_idx: pb["is_index"] = True
                            fin = process_table(page, t)
                            pb["content"].append({"type":"table","data":fin})
                            stats["tables"] += 1
                            logging.info(
                                f"Page {page.page_number} [pdfplumber]: "
                                f"{len(fin)}r × {len(fin[0]) if fin else 0}c ✓"
                            )
                    else:
                        lines = [l for l in text.split("\n") if l.strip()]
                        if lines: pb["content"].append({"type":"text","data":lines})

                elif pg_type == "h_only":
                    rows = extract_gap_based(page)
                    if rows:
                        pb["content"].append({"type":"table","data":rows})
                        stats["tables"] += 1
                        logging.info(
                            f"Page {page.page_number} [gap-based]: "
                            f"{len(rows)}r × {len(rows[0]) if rows else 0}c ✓"
                        )
                    else:
                        lines = [l for l in text.split("\n") if l.strip()]
                        if lines: pb["content"].append({"type":"text","data":lines})
                        logging.warning(f"Page {page.page_number} [gap-based]: fallback to text")

                else:  # no_lines / scanned
                    # Try gap-based first (fast), docling not available in browser
                    rows = extract_gap_based(page)
                    if rows:
                        pb["content"].append({"type":"table","data":rows})
                        stats["tables"] += 1
                        logging.info(
                            f"Page {page.page_number} [gap-fallback]: "
                            f"{len(rows)}r × {len(rows[0]) if rows else 0}c ✓"
                        )
                    else:
                        lines = [l for l in text.split("\n") if l.strip()]
                        if lines: pb["content"].append({"type":"text","data":lines})
                        logging.warning(
                            f"Page {page.page_number} [{pg_type}]: "
                            f"gap failed, stored as text. Install docling for full support."
                        )

                document.append(pb)

        # Build workbook
        wb, used = Workbook(), set()
        wb.remove(wb.active)

        for pb in document:
            ws = wb.create_sheet(title=sname(pb, used))
            cr = 1
            title = f"Page {pb['page_number']}"
            if pb["form_name"]: title += f"   |   {pb['form_name']}"
            hc = ws.cell(row=1, column=1, value=title)
            hc.font=FSH; hc.fill=FDB; hc.border=BSH; hc.alignment=AM
            cr = 2

            for item in pb["content"]:
                if item["type"] == "table":
                    td = item["data"]
                    if not td: continue
                    nc = max(len(r) for r in td)
                    for ri, row in enumerate(td):
                        row = list(row) + [""]*(nc-len(row))
                        srow(ws, cr, row, classify(row, ri, pb["is_index"]), nc)
                        cr += 1
                    cr += 1
                else:
                    for line in item["data"]:
                        c=ws.cell(row=cr,column=1,value=line); c.font=FN; c.alignment=AW
                        cr += 1
                    cr += 1

            set_widths(ws)
            ws.freeze_panes = "A2"
            stats["sheets"] += 1

        buf = io.BytesIO()
        wb.save(buf); buf.seek(0)
        stats["forms"] = sorted(stats["forms"])
        stats["engines"] = sorted(stats.get("engines", set()))
        logging.info(f"Done — {stats['sheets']} sheets, {stats['tables']} tables extracted")
        return buf.getvalue(), stats

    finally:
        os.unlink(tmp_path)


# ══════════════════════════════════════════════════════════════════════════════
# MAIN UI
# ══════════════════════════════════════════════════════════════════════════════

def main():
    _setup_logging()

    # ── Hero ─────────────────────────────────────────────────────────────────
    st.markdown("""
    <div class="hero">
        <div class="hero-icon">📊</div>
        <h1 class="hero-title">IRDAI PDF Extractor</h1>
        <p class="hero-sub">
            Upload any IRDAI public disclosure PDF and instantly get<br>
            a clean, multi-sheet Excel file — one sheet per page.
        </p>
    </div>
    <hr class="divider">
    """, unsafe_allow_html=True)

    # ── Upload ────────────────────────────────────────────────────────────────
    st.markdown('<p class="section-label">📁 Upload PDF</p>', unsafe_allow_html=True)

    uploaded = st.file_uploader(
        label="Upload IRDAI PDF",
        type=["pdf"],
        help="Single IRDAI public disclosure PDF (max 200 MB)",
        label_visibility="collapsed",
    )

    # File info pill — appears only when file is selected
    if uploaded:
        size_mb = uploaded.size / (1024 * 1024)
        st.markdown(f"""
        <div class="file-pill">
            <span class="file-pill-icon">📄</span>
            <div>
                <p class="file-pill-name">{uploaded.name}</p>
                <p class="file-pill-size">{size_mb:.2f} MB &nbsp;·&nbsp; PDF</p>
            </div>
        </div>
        """, unsafe_allow_html=True)

    st.markdown("<hr class='divider'>", unsafe_allow_html=True)

    # ── If no file — friendly prompt and stop ────────────────────────────────
    if not uploaded:
        st.markdown("""
        <div style="text-align:center; padding:32px 0 16px; color:#94A3B8;">
            <div style="font-size:2.2rem; margin-bottom:10px;">☝️</div>
            <p style="font-size:0.92rem; font-weight:600; color:#64748B; margin:0 0 4px;">
                Select a PDF above to get started
            </p>
            <p style="font-size:0.8rem; margin:0;">
                One file at a time · All FORM L-* types supported
            </p>
        </div>
        """, unsafe_allow_html=True)
        st.markdown('<div class="footer">IRDAI PDF Extractor · Supports all FORM L-* disclosure types</div>',
                    unsafe_allow_html=True)
        return

    # ── Output filename ───────────────────────────────────────────────────────
    st.markdown('<p class="section-label">⚙️ Output filename</p>', unsafe_allow_html=True)
    stem        = Path(uploaded.name).stem
    output_name = st.text_input(
        label="Output filename",
        value=f"{stem}_IRDAI.xlsx",
        label_visibility="collapsed",
    )
    if not output_name.lower().endswith(".xlsx"):
        output_name += ".xlsx"

    st.markdown("<hr class='divider'>", unsafe_allow_html=True)

    # ── Extract button ────────────────────────────────────────────────────────
    st.markdown('<p class="section-label">⚡ Extraction</p>', unsafe_allow_html=True)

    extract_btn = st.button(
        "⚡  Extract to Excel",
        type="primary",
        use_container_width=True,
    )

    prog_slot = st.empty()
    msg_slot  = st.empty()

    # ── Run ───────────────────────────────────────────────────────────────────
    if extract_btn:
        for k in ["excel_bytes", "stats", "logs"]:
            st.session_state.pop(k, None)

        prog = prog_slot.progress(0, text="Reading PDF…")
        msg_slot.info("Processing — this may take a moment for large files…", icon="⏳")

        try:
            pdf_bytes = uploaded.read()
            prog.progress(20, text="Detecting tables…")
            excel_bytes, stats = _run_extraction(pdf_bytes)
            prog.progress(90, text="Styling Excel…")
            time.sleep(0.3)
            prog.progress(100, text="Complete ✓")
            time.sleep(0.4)
            prog_slot.empty()
            msg_slot.empty()
            st.session_state["excel_bytes"] = excel_bytes
            st.session_state["stats"]       = stats

        except Exception as exc:
            prog_slot.empty()
            msg_slot.error(f"Extraction failed: {exc}", icon="🚨")
            return

    # ── Results ───────────────────────────────────────────────────────────────
    excel_bytes = st.session_state.get("excel_bytes")
    stats       = st.session_state.get("stats", {})

    if excel_bytes and stats:

        st.success("Extraction complete — your file is ready!", icon="✅")

        # Stats
        p = stats.get("pages",  0)
        t = stats.get("tables", 0)
        s = stats.get("sheets", 0)
        st.markdown(f"""
        <div class="stats-row">
            <div class="stat-card">
                <div class="stat-num">{p}</div>
                <div class="stat-lbl">Pages</div>
            </div>
            <div class="stat-card">
                <div class="stat-num">{t}</div>
                <div class="stat-lbl">Tables</div>
            </div>
            <div class="stat-card">
                <div class="stat-num">{s}</div>
                <div class="stat-lbl">Sheets</div>
            </div>
        </div>
        """, unsafe_allow_html=True)

        # Forms detected
        forms = stats.get("forms", [])
        if forms:
            badges = "".join(f'<span class="badge">{f}</span>' for f in forms)
            st.markdown(
                f'<p class="section-label">Forms detected</p>'
                f'<div class="badges">{badges}</div>',
                unsafe_allow_html=True,
            )

        st.markdown("<hr class='divider'>", unsafe_allow_html=True)

        # Download
        st.download_button(
            label="⬇️  Download Excel",
            data=excel_bytes,
            file_name=output_name,
            mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
            use_container_width=True,
        )

    # ── Log ───────────────────────────────────────────────────────────────────
    if st.session_state.get("logs"):
        st.markdown("<hr class='divider'>", unsafe_allow_html=True)
        st.markdown('<p class="section-label">📋 Processing log</p>', unsafe_allow_html=True)
        _render_logs()

    # ── Footer ────────────────────────────────────────────────────────────────
    st.markdown(
        '<div class="footer">IRDAI PDF Extractor · Supports all FORM L-* disclosure types</div>',
        unsafe_allow_html=True,
    )


if __name__ == "__main__":
    main()